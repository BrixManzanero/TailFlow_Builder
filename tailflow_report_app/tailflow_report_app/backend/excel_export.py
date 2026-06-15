"""Excel export: render the consolidated report data as a styled .xlsx.

Produces a Revenue (GMV & NMV) sheet that mirrors the on-screen report — navy
GMV band, green NMV band, a MoM column — plus one Product Performance sheet per
period (month or year) with the per-SKU change column. Totals and revenue MoM
use Excel formulas so the workbook stays live; product MoM is written as a
value (cross-period SKU lookups would be fragile as formulas).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

NAVY, GREEN, DARK = "1F3864", "2E7D46", "16233A"
WHITE, POS, NEG = "FFFFFF", "15803D", "C2453B"
STRIPE = "F4F7FA"
FONT = "Arial"
MONEY, PCT, INT = "#,##0.00", "0.0%", "#,##0"
MOM_FMT = "+0.0%;-0.0%"

_thin = Side(style="thin", color="D3DAE2")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
PLATFORMS = ["Shopee", "Shopify", "Lazada", "TikTok"]


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _header(cell, color, align="right"):
    cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
    cell.fill = _fill(color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER


def _data(cell, fmt=None, bold=False, color=None, align="right"):
    cell.font = Font(name=FONT, bold=bold, color=color or "000000")
    cell.alignment = Alignment(horizontal=align)
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def _mom_cf(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
                                                  font=Font(name=FONT, color=NEG)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
                                                  font=Font(name=FONT, color=POS)))


def _revenue_sheet(ws, rev):
    P = PLATFORMS
    ws.merge_cells("A1:A2")
    _header(ws["A1"], NAVY, "left"); ws["A1"] = "Month"
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    _header(ws.cell(1, 2, "GMV by Platform (PHP)"), NAVY, "center")
    ws.merge_cells("G1:G2")
    _header(ws.cell(1, 7, "MoM Δ%"), DARK, "center")
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
    _header(ws.cell(1, 8, "NMV by Platform (PHP)"), GREEN, "center")
    for i, h in enumerate(P + ["Total GMV"]):
        _header(ws.cell(2, 2 + i, h), NAVY)
    for i, h in enumerate(P + ["Total NMV"]):
        _header(ws.cell(2, 8 + i, h), GREEN)

    r0 = 3
    for idx, row in enumerate(rev):
        r = r0 + idx
        stripe = STRIPE if idx % 2 else WHITE
        mc = ws.cell(r, 1, row["month"])
        _data(mc, bold=True, align="left"); mc.fill = _fill("EEF1F5")
        for i, p in enumerate(P):
            c = ws.cell(r, 2 + i, row["gmv"][p]); _data(c, MONEY); c.fill = _fill(stripe)
        tg = ws.cell(r, 6, f"=SUM(B{r}:E{r})"); _data(tg, MONEY, bold=True); tg.fill = _fill("EAF0F7")
        mm = ws.cell(r, 7)
        if idx > 0:
            mm.value = f'=IFERROR((F{r}-F{r-1})/F{r-1},"")'
            _data(mm, MOM_FMT, bold=True)
        else:
            _data(mm)
        for i, p in enumerate(P):
            c = ws.cell(r, 8 + i, row["nmv"][p]); _data(c, MONEY); c.fill = _fill(stripe)
        tn = ws.cell(r, 12, f"=SUM(H{r}:K{r})"); _data(tn, MONEY, bold=True); tn.fill = _fill("E9F4ED")

    last = r0 + len(rev) - 1
    tr = last + 1
    tc = ws.cell(tr, 1, "TOTAL"); _data(tc, bold=True, color=WHITE, align="left"); tc.fill = _fill(DARK)
    for col in [2, 3, 4, 5, 6, 8, 9, 10, 11, 12]:
        L = get_column_letter(col)
        c = ws.cell(tr, col, f"=SUM({L}{r0}:{L}{last})")
        _data(c, MONEY, bold=True, color=WHITE); c.fill = _fill(DARK)
    g7 = ws.cell(tr, 7); _data(g7, bold=True, color=WHITE); g7.fill = _fill(DARK)

    _mom_cf(ws, f"G{r0}:G{last}")
    ws.column_dimensions["A"].width = 11
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "B3"


def _product_sheet(ws, period, blk, kind):
    chg_label = "MoM Δ%" if kind == "monthly" else "YoY Δ%"
    tag = "MoM" if kind == "monthly" else "YoY"
    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, f"Product Performance ({tag}) — {period}")
    t.font = Font(name=FONT, bold=True, color=WHITE, size=12)
    t.fill = _fill(GREEN); t.alignment = Alignment(horizontal="left", vertical="center")
    heads = ["#", "Product", "SKU", "GMV (PHP)", chg_label, "Qty Sold", "Stock", "Sell-Through"]
    aligns = ["center", "left", "left", "right", "right", "right", "right", "right"]
    for i, (h, a) in enumerate(zip(heads, aligns)):
        _header(ws.cell(2, 1 + i, h), GREEN, a)

    r0 = 3
    rows = blk["rows"]
    for idx, p in enumerate(rows):
        r = r0 + idx
        stripe = STRIPE if idx % 2 else WHITE
        for col in range(1, 9):
            ws.cell(r, col).fill = _fill(stripe)
        _data(ws.cell(r, 1, idx + 1))
        _data(ws.cell(r, 2, p["name"]), align="left")
        _data(ws.cell(r, 3, p["sku"]), align="left", color="5A6675")
        _data(ws.cell(r, 4, p["gmv"]), MONEY)
        cc = ws.cell(r, 5)
        if p["chg"] is None:
            cc.value = "—"; _data(cc)
        else:
            cc.value = p["chg"] / 100
            _data(cc, MOM_FMT, color=POS if p["chg"] >= 0 else NEG)
        _data(ws.cell(r, 6, p["qty"]), INT)
        _data(ws.cell(r, 7, p["stock"]), INT)
        _data(ws.cell(r, 8, p["st"]), PCT)

    last = r0 + len(rows) - 1
    tr = last + 1
    for col in range(1, 9):
        ws.cell(tr, col).fill = _fill(DARK)
    lab = ws.cell(tr, 2, f"TOTAL · {blk['totals']['count']} products")
    _data(lab, bold=True, color=WHITE, align="left")
    _data(ws.cell(tr, 4, f"=SUM(D{r0}:D{last})"), MONEY, bold=True, color=WHITE)
    _data(ws.cell(tr, 6, f"=SUM(F{r0}:F{last})"), INT, bold=True, color=WHITE)
    _data(ws.cell(tr, 7, f"=SUM(G{r0}:G{last})"), INT, bold=True, color=WHITE)
    for col in (1, 3, 5, 8):
        _data(ws.cell(tr, col), bold=True, color=WHITE)

    for col, w in zip("ABCDEFGH", [5, 42, 12, 16, 11, 11, 10, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


def build_excel(data):
    wb = Workbook()
    wb.remove(wb.active)
    if data.get("revenue"):
        _revenue_sheet(wb.create_sheet("Revenue (GMV & NMV)"), data["revenue"])
    for year, blk in (data.get("products_yearly") or {}).items():
        _product_sheet(wb.create_sheet(f"YoY {year}"[:31]), year, blk, "yearly")
    for month, blk in (data.get("products_monthly") or {}).items():
        _product_sheet(wb.create_sheet(f"MoM {month}"[:31]), month, blk, "monthly")
    if not wb.sheetnames:
        ws = wb.create_sheet("Report")
        ws["A1"] = "Upload exports to populate this report."
    try:
        wb.calculation.fullCalcOnLoad = True  # force recalc when opened
    except Exception:
        pass
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
